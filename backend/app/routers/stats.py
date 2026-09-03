"""
Excel'dan yuklangan haqiqiy statistika endpointlari.

Eski `/api/analytics/*` reja↔amalda demo ma'lumoti bilan ishlaydi va
admin paneli uchun qoladi. Bosh sahifa esa shu yerdagi `/api/stats/*`
dan oziqlanadi — 1084 ko'rsatkich, 24 199 o'lchov, 2010–2026 yillar.
"""

import io
import re
from datetime import datetime

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.database import get_db
from app.ingest.excel import Record, _YEAR_MAX, _YEAR_MIN, parse_bytes
from app.ingest.loader import CATEGORIES, keep_known, load_records
from app.models import District, StatCategory, StatIndicator, StatObservation
from app.security import require_admin
from app.services import stats as st

router = APIRouter(prefix="/api/stats", tags=["stats"])


def _resolve(
    db: Session, indicator_id: int | None, module: str | None
) -> StatIndicator:
    ind = st.resolve_indicator(db, indicator_id=indicator_id, module=module)
    if ind is None:
        raise HTTPException(404, "Kórsetkish tabılmadı")
    return ind


def _year(db: Session, year: int | None, indicator_id: int | None = None) -> int:
    return year or st.latest_year(db, indicator_id)


@router.get("/meta")
def meta(db: Session = Depends(get_db)):
    """Boshlang'ich ma'lumot: yillar, sohalar, kategoriyalar, hududlar."""
    return st.meta(db)


@router.get("/overview")
def overview(year: int | None = None, db: Session = Depends(get_db)):
    return st.overview(db, _year(db, year))


@router.get("/map")
def map_layer(
    module: str | None = None,
    indicator_id: int | None = None,
    year: int | None = None,
    db: Session = Depends(get_db),
):
    """Xarita qatlami — tuman kesimidagi qiymatlar va rang jadali."""
    ind = _resolve(db, indicator_id, module)
    if not ind.has_districts:
        raise HTTPException(400, "Bul kórsetkishte rayonlar kesimi joq")
    return st.map_layer(db, ind, _year(db, year, ind.id))


@router.get("/series")
def series(
    module: str | None = None,
    indicator_id: int | None = None,
    district_id: str | None = None,
    year_from: int | None = None,
    year_to: int | None = None,
    db: Session = Depends(get_db),
):
    """Yillar kesimidagi qator (2010–2026)."""
    ind = _resolve(db, indicator_id, module)
    return {
        "indicator": st.indicator_brief(ind),
        "district_id": district_id,
        "unit": ind.unit,
        "points": st.series(
            db, ind, district_id=district_id, year_from=year_from, year_to=year_to
        ),
    }


@router.get("/districts/{district_id}")
def district_profile(
    district_id: str, year: int | None = None, db: Session = Depends(get_db)
):
    profile = st.district_profile(db, district_id, _year(db, year))
    if profile is None:
        raise HTTPException(404, "Aymaq tabılmadı")
    return profile


@router.get("/categories")
def categories(db: Session = Depends(get_db)):
    return st.meta(db)["categories"]


@router.get("/indicators")
def indicators(
    q: str | None = None,
    category_id: str | None = None,
    module: str | None = None,
    has_districts: bool | None = None,
    limit: int = Query(default=50, le=500),
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """Ko'rsatkichlar ma'lumotnomasi — qidiruv va sahifalash bilan."""
    return st.search_indicators(
        db,
        q=q,
        category_id=category_id,
        module=module,
        has_districts=has_districts,
        limit=limit,
        offset=offset,
    )


@router.get("/indicators/{indicator_id}")
def indicator_detail(indicator_id: int, db: Session = Depends(get_db)):
    return st.indicator_detail(db, _resolve(db, indicator_id, None))


@router.get("/indicators/{indicator_id}/breakdown")
def indicator_period_breakdown(
    indicator_id: int,
    year: int,
    district_id: str | None = None,
    db: Session = Depends(get_db),
):
    """Berilgen jıldaǵı barlıq dáwir kesimleri (jıllıq, yarım jıllıq, sherek, ay)."""
    ind = _resolve(db, indicator_id, None)
    return {
        "indicator": st.indicator_brief(ind),
        "year": year,
        "district_id": district_id,
        "unit": ind.unit,
        "points": st.period_breakdown(db, ind, year, district_id=district_id),
    }


# ── Admin ────────────────────────────────────────────────────────────


@router.get("/summary", dependencies=[Depends(require_admin)])
def summary(db: Session = Depends(get_db)):
    """Admin paneli uchun bazaning holati."""
    per_category = db.execute(
        select(
            StatCategory.id,
            StatCategory.name_kaa,
            StatCategory.source_dir,
            StatCategory.color,
            func.count(StatIndicator.id),
        )
        .outerjoin(StatIndicator, StatIndicator.category_id == StatCategory.id)
        .group_by(StatCategory.id)
        .order_by(StatCategory.sort)
    ).all()

    observations = dict(
        db.execute(
            select(StatIndicator.category_id, func.count(StatObservation.id))
            .join(StatObservation, StatObservation.indicator_id == StatIndicator.id)
            .group_by(StatIndicator.category_id)
        ).all()
    )

    years = st.available_years(db)
    return {
        "years": years,
        "latest_year": years[-1] if years else None,
        "indicators": db.scalar(select(func.count()).select_from(StatIndicator)) or 0,
        "observations": db.scalar(select(func.count()).select_from(StatObservation)) or 0,
        "districts": db.scalar(
            select(func.count(func.distinct(StatObservation.district_id)))
        ) or 0,
        "with_districts": db.scalar(
            select(func.count()).select_from(StatIndicator).where(
                StatIndicator.has_districts.is_(True)
            )
        ) or 0,
        "categories": [
            {
                "id": cid,
                "name": name,
                "source_dir": source_dir,
                "color": color,
                "indicators": n,
                "observations": observations.get(cid, 0),
            }
            for cid, name, source_dir, color, n in per_category
        ],
        #: Yuklash formasidagi tanlov — manba papkalari
        "source_dirs": sorted(CATEGORIES),
        #: Har bir tayanch soha uchun hozir qaysi ko'rsatkich ishlatilyapti
        "modules": [
            {
                "id": module,
                "name": st.MODULE_META[module][0],
                "color": st.MODULE_META[module][2],
                "indicator_id": ind.id,
                "indicator_name": ind.name_kaa,
                "unit": ind.unit,
            }
            for module, ind in st.primary_indicators(db).items()
            if module in st.MODULE_META
        ],
    }


class IndicatorPatch(BaseModel):
    """Ko'rsatkichni tayanch soha sifatida belgilash yoki bo'shatish."""

    #: `sanaat`, `awil_xojaligi`, ... yoki bo'sh satr — biriktirishni uzish
    module: str | None = None
    lower_is_better: bool | None = None


class IndicatorBulkPatch(BaseModel):
    """Bir nechte ko'rsatkichti birdey tayanch sohaǵa biriktiriw."""

    ids: list[int]
    module: str | None = None


@router.patch("/indicators/bulk", dependencies=[Depends(require_admin)])
def bulk_update_indicators(payload: IndicatorBulkPatch, db: Session = Depends(get_db)):
    """
    Bir nechew qatardı birden biriktiredi.

    DIQQAT: bul marshrut `/indicators/{indicator_id}` den ALDIN turıwı
    SHÁRT — Starlette jol salıstırıwdı registraciya tártibinde islaydi,
    aks halda "bulk" sózi `indicator_id` retinde islenip, 422 qátelik
    beredi.
    """
    module = (payload.module or "").strip() or None
    if module is not None and module not in st.MODULE_META:
        raise HTTPException(400, f"Belgisiz taraw: {module}")

    updated, skipped = 0, 0
    found = {
        ind.id: ind
        for ind in db.scalars(
            select(StatIndicator).where(StatIndicator.id.in_(payload.ids))
        )
    }
    for indicator_id in payload.ids:
        ind = found.get(indicator_id)
        # Rayon kesimisiz ko'rsatkichke tayanch soha biriktirilmeydi —
        # jalǵız-jalǵız PATCH'tegi tekseriwdiń ózi, tek jańlısı ótkerip
        # jiberiledi, butın ámel toqtamaydı
        if ind is None or (module is not None and not ind.has_districts):
            skipped += 1
            continue
        ind.module = module
        updated += 1

    db.commit()
    return {"updated": updated, "skipped": skipped}


@router.patch("/indicators/{indicator_id}", dependencies=[Depends(require_admin)])
def update_indicator(
    indicator_id: int, payload: IndicatorPatch, db: Session = Depends(get_db)
):
    ind = db.get(StatIndicator, indicator_id)
    if ind is None:
        raise HTTPException(404, "Kórsetkish tabılmadı")

    fields = payload.model_dump(exclude_unset=True)
    if "module" in fields:
        module = (fields["module"] or "").strip() or None
        if module is not None and module not in st.MODULE_META:
            raise HTTPException(400, f"Belgisiz taraw: {module}")
        if module is not None and not ind.has_districts:
            # Tayanch ko'rsatkich xaritani bo'yaydi — rayon kesimisiz
            # xarita bo'sh qoladi, shuning uchun bunday biriktirish rad etiladi
            raise HTTPException(400, "Bul kórsetkishte rayonlar kesimi joq")
        ind.module = module
    if "lower_is_better" in fields:
        ind.lower_is_better = bool(fields["lower_is_better"])

    db.commit()
    db.refresh(ind)
    return st.indicator_detail(db, ind)


async def _parse_upload(file: UploadFile, category: str) -> list[Record]:
    """Yuklangan faylni yozuvlar oqimiga aylantiradi — preview va haqiqiy yuklashda umumiy."""
    if category not in CATEGORIES:
        raise HTTPException(400, f"Belgisiz bólim: {category}")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Tek ǵana .xlsx yamasa .xls fayl qabıl etiledi")

    raw = await file.read()
    try:
        records = keep_known(
            parse_bytes(raw, category=category, filename=file.filename or "upload.xlsx")
        )
    except Exception as exc:  # noqa: BLE001 — buzuq fayl 500 bermasin
        raise HTTPException(400, f"Faydı oqıp bolmadı: {exc}") from exc

    if not records:
        raise HTTPException(
            400,
            "Fayldan bir de jazıw alınbadı — dáwir atları bar sarlawha qatarı tabılmadı",
        )
    return records


def _sample_records(records: list[Record], db: Session, limit: int = 20) -> list[dict]:
    """
    Faylda parsıńǵannan bir nesheu haqıyqıy qatar — admin sanlarǵa ǵana emes,
    NAMA oqılǵanına da kóz jetkiziwi ushın (rayon durıs tanılǵanba, san
    durıs shıǵıp turba). Hár kórsetkishten bir qatar, kóp bolsa `limit`ge
    shekem — sonda bir ǵana kórsetkishtiń on qatarı emes, alwan túrlilik
    kórinedi.
    """
    dists = st.district_names(db)
    by_indicator: dict[str, Record] = {}
    for r in records:
        by_indicator.setdefault(r.indicator, r)
    return [
        {
            "korsetkish": r.indicator,
            "rayon": dists[r.district_id].name if r.district_id in dists else "Respublika",
            "jıl": r.year,
            "dawir": st.period_label(r.period, r.period_no),
            "qıymet": r.value,
            "olshem": r.unit,
        }
        for r in list(by_indicator.values())[:limit]
    ]


def _build_upload_template(db: Session) -> io.BytesIO:
    """
    Júklew ushın úlgi — eń keń tarqalǵan "dáwirler bağаnada, rayonlar
    qatarda" forması (`ingest/excel.py`dağı `_parse_wide`). "Úlgi"
    varağında rayon atları HAQIYQIY (bazadan alınǵan, sonlıqtan
    tanılıwı kepillengen) hám mısal sanlar menen tayar keste bar — usı
    kúyinde de "Ko'riw" túymesi arqalı dárhal parslanadı, sonda admin
    format durıslığın júklewden aldın kóre aladı.
    """
    districts = db.scalars(select(District).order_by(District.name)).all()
    year0 = datetime.now().year
    years = [year0 - 2, year0 - 1, year0]

    wb = Workbook()

    info = wb.active
    info.title = "Nusqama"
    info.sheet_view.showGridLines = False
    info.column_dimensions["A"].width = 92
    lines = [
        ("Bul fayldı qalay tolтırıw kerek", True),
        ("", False),
        ("«Úlgi» varağında bir kórsetkishtiń mısalı bar — rayonlar qatarda, jıllar bağаnada.", False),
        ("1-qatardağı sarlawha atın (A1) óz kórsetkishińiздиń atına ózgertiń.", False),
        ("«Ólshem birligi» bağаnasına birlikti jazıń — mısalı: mlrd. som, %, adam, dana, ga.", False),
        (f"Jıl bağаnaların kerekli sanğa deyin qosıń yamasa ózgertiń ({_YEAR_MIN}–{_YEAR_MAX} aralığında).", False),
        ("Rayon atların ÓZGERTPEŃ — dál usı jazılıwda tanıladı; qatar tártibi áhmiyetsiz.", False),
        ("Belgisiz yamasa boS qaldırılğan katak — sol dáwir ushın maǵlıwmat joq dep esaplanadı.", False),
        ("Bir fayldıń bir varağında bir ǵana kórsetkish bolsa — eń isenimli nátiyje sonda.", False),
        ("", False),
        ("Tolтırıp bolğannan keyin: admin panelde «Bólim» tańlań, faylды taslań, «Ko'riw»", False),
        ("túymesin basıń — server namuna qatarlardı kórsetedi, sol jerde tekserip alasız.", False),
        ("Hesh nárse saqlanbaydı, «Tastıyıqlaw hám bazağa jazıw» basılğanşa.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = info.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=13 if bold else 11, color="1F2937")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet = wb.create_sheet("Úlgi")
    sheet.sheet_view.showGridLines = False
    columns = ["Rayon", "Ólshem birligi", *[str(y) for y in years]]

    # 1-qatar — kórsetkish atı (parser bunı "sarlawha" dep oqıydı, dáwir
    # qatarınıń ÚSTİNDE turıwı shárt); 2-qatar — dáwir sarlawhaları;
    # 3-qatardan maǵlıwmat baslanadı.
    title_cell = sheet.cell(row=1, column=1, value="Ónim kólemi (mısal — óz kórsetkishińizdiń atına ózgertiń)")
    title_cell.font = Font(bold=True, size=12, color="1F2937")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    header_row = 2
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_border = Border(bottom=Side(style="thin", color="1E3A8A"))
    for idx, name in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(vertical="center")

    for r, d in enumerate(districts, start=header_row + 1):
        sheet.cell(row=r, column=1, value=d.name)
        sheet.cell(row=r, column=2, value="mlrd. som")
        for c, _year in enumerate(years, start=3):
            sheet.cell(row=r, column=c, value=round(100 + r * 3.7 + c * 5.1, 1))

    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 16
    for c in range(3, 3 + len(years)):
        sheet.column_dimensions[get_column_letter(c)].width = 12
    sheet.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/upload/template", dependencies=[Depends(require_admin)])
def upload_template(db: Session = Depends(get_db)):
    """Tolтırılıwı kerek úlgi Excel fayldı qaytaradı (admin panelindegi «Shablon» túymesi)."""
    buf = _build_upload_template(db)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="statistika-shablon.xlsx"'},
    )


@router.post("/upload/preview", dependencies=[Depends(require_admin)])
async def upload_preview(
    file: UploadFile = File(...),
    category: str = Form(...),
    db: Session = Depends(get_db),
):
    """
    Fayldı bazaǵa jazbastan sınap kóredi.

    Nátiyje `/upload` menen bir túrde — sonday kórsetkishler jańalanadı,
    sonday sanlar shıǵadı — biraq tranzaksiya HÁMİSHE keri qaytarıladı
    (`db.rollback()`), sonlıqtan baza tiimeydi. Admin nadurıs bólim
    tańlaǵanın júklewden ALDIN kóredi.
    """
    records = await _parse_upload(file, category)
    try:
        stats = load_records(records, db, replace_all=False, commit=False)
    finally:
        db.rollback()
    return {
        "file": file.filename,
        "category": category,
        "preview": True,
        "namuna": _sample_records(records, db),
        **stats,
    }


@router.post("/upload", dependencies=[Depends(require_admin)])
async def upload_workbook(
    file: UploadFile = File(...),
    category: str = Form(...),
    db: Session = Depends(get_db),
):
    """
    Statistika Excel faylini bazaga yuklaydi.

    Fayl qaysi bo'limga tegishli ekani nomidan bilinmaydi (manba
    papkalarida turgan), shuning uchun `category` ochiq beriladi.
    Yuklash faqat SHU fayl tegib o'tgan ko'rsatkichlarni yangilaydi.
    """
    records = await _parse_upload(file, category)
    stats = load_records(records, db, replace_all=False)
    return {"file": file.filename, "category": category, "preview": False, **stats}


_EXPORT_COLUMNS = [
    "Bólim", "Kórsetkish", "Ólshem", "Rayon", "Jıl",
    "Dáwir", "Dáwir nomeri", "Qıymet", "Reja", "Derek",
]
_EXPORT_NUMBER_COLUMNS = {"Qıymet", "Reja"}
_SHEET_NAME_BAD = re.compile(r"[\\/*?:\[\]]")


def _sheet_name(label: str, taken: set[str]) -> str:
    """
    Excel varaq atı — eń kópi 31 belgi, `\\/*?:[]` bolmawı kerek.

    Bir neshe bólim qısqartıwdan keyin birdey nomǵa túsiwi múmkin
    (mısalı eki uzın nomniń birinshi 31 belgisi sáykes keliwi); bunday
    jaǵdayda san qosıp ajıratamız, bolmasa `openpyxl` qátelik beredi.
    """
    base = _SHEET_NAME_BAD.sub("_", label).strip()[:31] or "Bólim"
    name, n = base, 2
    while name in taken:
        suffix = f" ({n})"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    taken.add(name)
    return name


def _write_sheet(ws, df: pd.DataFrame, title: str) -> None:
    """Bitta varaqqa jazıw: title qatarı, formatlanǵan header, keńlikler, freeze."""
    n_cols = len(_EXPORT_COLUMNS)
    header_row = 2
    first_data_row = header_row + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    subtitle = f"{len(df)} qatar · {datetime.now():%Y-%m-%d %H:%M}"
    title_cell = ws.cell(row=1, column=1, value=f"{title} · {subtitle}")
    title_cell.font = Font(bold=True, size=12, color="1F2937")
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_border = Border(bottom=Side(style="thin", color="1E3A8A"))
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(vertical="center")

    for idx, name in enumerate(_EXPORT_COLUMNS, start=1):
        col_letter = get_column_letter(idx)
        content_width = df[name].astype(str).map(len).max() if len(df) else 0
        width = max(len(name), content_width) + 2
        ws.column_dimensions[col_letter].width = min(max(width, 8), 42)
        if name in _EXPORT_NUMBER_COLUMNS:
            for cell in ws[col_letter][first_data_row - 1:]:
                cell.number_format = "#,##0.###"

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{ws.max_row}"
    ws.print_title_rows = f"{header_row}:{header_row}"


@router.get("/export", dependencies=[Depends(require_admin)])
def export_data(
    category_id: str | None = None,
    db: Session = Depends(get_db),
):
    """
    Bazadagi o'lchovlarni Excel qilip qaytaradi.

    `category_id` berilse — bir varaqlı fayl, tek sol bólim. Berilmese —
    butun baza, hár bólim ÓZ VARAǴINDA (bir tegis 24 mıń qatarlı varaqta
    bólimler bir-birine sińip, kóz menen aylanıw qıyın bolar edi).
    24 mıńǵa jaqın qatar hátte kóp varaqlı xlsx'te de tez jaratıladı,
    sonlıqtan fon jumısı kerek emes.
    """
    stmt = (
        select(
            StatCategory.name_kaa,
            StatIndicator.name_kaa,
            StatIndicator.unit,
            District.name,
            StatObservation.year,
            StatObservation.period,
            StatObservation.period_no,
            StatObservation.value,
            StatObservation.plan_value,
            StatIndicator.source,
        )
        .select_from(StatObservation)
        .join(StatIndicator, StatObservation.indicator_id == StatIndicator.id)
        .join(StatCategory, StatIndicator.category_id == StatCategory.id)
        .outerjoin(District, StatObservation.district_id == District.id)
        .order_by(StatCategory.sort, StatIndicator.name_kaa, StatObservation.year)
    )
    category_name = None
    if category_id:
        category_name = db.scalar(
            select(StatCategory.name_kaa).where(StatCategory.id == category_id)
        )
        if category_name is None:
            raise HTTPException(400, f"Belgisiz bólim: {category_id}")
        stmt = stmt.where(StatIndicator.category_id == category_id)

    rows = db.execute(stmt).all()
    df = pd.DataFrame(rows, columns=_EXPORT_COLUMNS)
    # Bos "Rayon" — respublika boyınsha jıyındı qıymet, hudud emes
    df["Rayon"] = df["Rayon"].fillna("Respublika")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if category_id or df.empty:
            title = f"Qaraqalpaqstan statistikası — {category_name or 'barlıq bólimler'}"
            df.to_excel(writer, index=False, sheet_name="Statistika", startrow=1)
            _write_sheet(writer.sheets["Statistika"], df, title)
        else:
            taken: set[str] = set()
            # `sort=False` — qatarlar db'den bólim tártibi (`StatCategory.sort`)
            # boyınsha kelgen, groupby usı tártipti buzbasın
            for label, part in df.groupby("Bólim", sort=False):
                sheet = _sheet_name(str(label), taken)
                part.to_excel(writer, index=False, sheet_name=sheet, startrow=1)
                _write_sheet(writer.sheets[sheet], part, f"Qaraqalpaqstan statistikası — {label}")

    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    buf.seek(0)

    # Content-Disposition sarlawhasın buzbaw ushın — tek qáwipsiz belgiler
    safe_id = re.sub(r"[^\w-]", "_", category_id) if category_id else "barlik"
    filename = f"statistika-{safe_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
